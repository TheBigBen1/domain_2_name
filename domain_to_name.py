# -*- coding: utf-8 -*-
"""
Created on Wed May  8 01:00:09 2019

@author: Benjamin
"""

import requests
from bs4 import BeautifulSoup
import html2text
from selenium import webdriver
import threading
import queue
import difflib
from tld import get_fld, get_tld
import urllib
from lxml import html
from lxml.html.clean import clean_html
import string
from urllib.parse import urlparse
from cleanco import cleanco, legal_suffixes, eng_words
import re 
import queue
import time
from selenium import webdriver
import re
import html2text
import spacy
nlp = spacy.load('en')
import openpyxl
from urllib.error import URLError
from urllib.request import Request

User_ag='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36'

h = html2text.HTML2Text()
h.body_width = 0
h.ignore_links = True
h.ignore_images = True

"""
why the fuck does headless selenium not redirect for www.bankofscotland.co.uk,
(only without headless)

- try to recognize text from logo (use clearbit logo api)
- Search on site for div with "logo" in name or ID and crawl ALT value
"""






imprint_trigger_words = ['imprint', 'legal',"terms","cookie",
                         'afdruk', 'بصمة', 'imprint', 'адбітак', 'отпечатък', 'অঙ্কিত করা', 'impremta', 'imprint', 'otisk', 'argraffnod', 'aftryk', 'impressum', 'αποτύπωμα', 'imprint', 'imprint', 'imprimir', 'jäljend', 'aztarna', 'اثر', 'jälki', 'imprimer', 'rian', 'impresión', 'છાપ', 'shafi', 'छाप', 'imprint', 'otisak', 'anprint', 'impresszum', 'տպագիր', 'jejak', 'akara', 'áletrun', 'impronta', 'טביעת רגל', '刻印', 'imprint', 'imprint', 'ізі', 'បោះពុម្ព', 'ಮುದ್ರೆ', '날인', 'imprint', 'imprint', 'įspaudas', 'nospiedums', 'marika', 'tuhinga', 'отпечаток', 'മുദ്രണം ചെയ്യുക', 'дардас', 'छाप', 'jejak', 'stampat', 'နှိပ်ရာ', 'छाप्नुहोस्', 'afdruk', 'avtrykk',
                     'zolemba', 'ਛਾਪ', 'odcisk', 'imprimir', 'imprima', 'отпечаток', 'මුද්රණය', 'odtlačok', 'odtis', 'riix', 'imprint', 'импринт', 'mongolo', 'kaluaran', 'avtryck', 'alama', 'முத்திரையில்', 'ముద్రణ', 'данд', 'ประทับ', 'imprint', 'damga', 'відбиток', 'امپرنٹ', 'bosma', 'dấu ấn', 'אָפּדרוק', 'isamisi', '版本说明', '版本说明', '版本說明', 'imprint', 'vrywaring', 'تنصل', 'disclaimer', 'адмова', 'опровержение', 'দাবি পরিত্যাগী', 'disclaimer', 'renúncia de responsabilitat', 'pagsaway', 'vyloučení odpovědnosti', 'ymwadiad', 'ansvarsfraskrivelse', 'haftungsausschluss', 'αποποίηση ευθυνών', 'disclaimer', 'disclaimer', 'renuncia', 'vastutusest loobumine',
                     'disclaimer', 'سلب مسئولیت', 'vastuuvapauslauseke', 'avertissement', 'séanadh', 'exención de responsabilidade', 'ડિસક્લેમર', 'disclaimer', 'अस्वीकरण', 'disclaimer', 'odricanje', 'avètisman', 'lemondás', 'հերքում', 'penolakan', 'nkwenye', 'fyrirvari', 'disconoscimento', 'כתב ויתור', '免責事項', 'disclaimer', 'პასუხისმგებლობის უარყოფა', 'бас тарту туралы ескерту', 'បដិសេធ', 'ಹಕ್ಕು ನಿರಾಕರಣೆ', '기권', 'disclaimer', 'ຄໍາປະຕິເສດ', 'atsisakymas', 'atruna', 'disclaimer', 'tuhinga o mua', 'одрекување од одговорност', 'നിരാകരണം', 'татгалзах', 'अस्वीकरण', 'penafian', 'ċaħda', 'မသက်ဆိုင်ကြောင်းရှင်းလင်းချက်', 'अस्वीकरण', 'ontkenning', 'ansvarsfraskrivelse',
                     'chotsutsa', 'ਬੇਦਾਅਵਾ', 'zrzeczenie się', 'aviso legal', 'act de renunțare', 'отказ', 'විරහිත කිරීම', 'dementi', 'zavrnitev odgovornosti', 'diidmada', 'mohim', 'дисцлаимер', 'ho hloka boikemelo', 'bantahan', 'varning', 'kizuizi', 'பொறுப்பாகாமை', 'డిస్క్లైమర్', 'рад', 'คำปฏิเสธ', 'disclaimer', 'feragat', 'відмова від відповідальності', 'ڈس کلیمر', 'voz kechish', 'từ chối trách nhiệm', 'disclaimer', 'ṣe idaniloju', '放弃', '放弃', '放棄', 'ukuxoshwa', 'privaatheid', 'الإجمالية', 'gizlilik', 'канфідэнцыяльнасць', 'поверителност', 'গোপনীয়তা', 'privatnost', 'privadesa', 'privacy', 'soukromí', 'preifatrwydd', 'privatliv', 'privatsphäre', 'μυστικότητα',
                     'privacy', 'privacy', 'intimidad', 'privaatsust', 'pribatutasun', 'حریم خصوصی', 'yksityisyys', 'intimité', 'príobháideacht', 'privacidade', 'ગોપનીયતા', 'sirri', 'एकांत', 'privacy', 'privatnost', 'vi prive', 'magánélet', 'գաղտնիությունը', 'pribadi', 'nzuzo', 'næði', 'privacy', 'פרטיות', 'プライバシー', 'privasi', 'კონფიდენციალურობა', 'құпиялылық', 'ភាពឯកជន', 'ಗೌಪ್ಯತೆ', '은둔', 'privacy', 'ຄວາມເປັນສ່ວນຕົວ', 'privatumas', 'privātums', 'ny fiainana manokana', 'tūmataiti', 'приватност', 'സ്വകാര്യത', 'хувийн нууцлал', 'गोपनीयता', 'privasi', 'privatezza', 'privacy ကို', 'गोपनीयता', 'privacy', 'personvern', 'zachinsinsi', 'ਗੋਪਨੀਯਤਾ', 'prywatność',
                     'privacidade', 'intimitate', 'конфиденциальность', 'පෞද්ගලිකත්වය', 'súkromia', 'zasebnost', 'asturnaanta', 'intimitet', 'приватност', 'sephiri', 'kalaluasaan', 'integritet', 'faragha', 'தனியுரிமை', 'గోప్యతా', 'махфият', 'ความเป็นส่วนตัว', 'privacy', 'gizlilik', 'конфіденційність', 'رازداری', 'maxfiylik', 'riêng tư', 'פּריוואַטקייט', 'ìpamọ', '隐私', '隐私', '隱私', 'ubumfihlo',
                     'agb', "datenschutz"]
# Also translate "Alle rechte vorbehalten "

hdr = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36',
   'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
   'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
   'Accept-Encoding': 'none',
   'Accept-Language': 'en-US,en;q=0.8',
   'Connection': 'keep-alive'
   }
copyright_replace = ["â©","©","â©","â "," .","by","all rights reserved","all right reserved","copyright"]

opts = webdriver.ChromeOptions()
opts.add_argument('headless')
driver = webdriver.Chrome(chrome_options=opts)
driver.implicitly_wait(15)
driver.set_page_load_timeout(15)


def name_cleaner(text, legal_clean=False):
    if text is None: return None
    if text=="": return None
    text = text.lower()
    for repl in copyright_replace: text = text.replace(repl,"")
    to_del = re.findall('(\d{4})',text)
    if len(to_del)>1: 
        text = text.replace(to_del[0]+"-"+to_del[1]," ")
        text = text.replace(to_del[0]+" - "+to_del[1]," ")
    for year in to_del: text = text.replace(year," ")
    if len(str(text).strip())<2: return None
    while True:
        if text[-1]=="." or text[-1]==" ": text = text[:-1]
        else: break
    text = re.sub(' +', ' ',text).lower().strip()
    text = re.sub(r'^[\.?&|*]', '',text)
    text = re.sub(r'[\.?&|*]$', '',text)
    if "(" in text:
        if text.find("(")<text.find(")"):
            text = text.replace(text[text.find("("):text.find(")")]+")"," ")
    if text[0]==".": text = text[1:]
    if legal_clean: text = cleanco(text).clean_name()
    return string.capwords(text.strip())
                
def copyright_crawler(tree):
    footer_tree = tree.xpath('//footer//text()') #alt: tree.xpath('//footer//*')
    copyright_element = []
    for element in footer_tree:
        el = element
        if el is None: continue
        el = el.replace("\n","").replace("\r","")
        el = re.sub(' +', ' ',el).lower()
        if el=="None" or el=="" or el==" ": continue
        if "©" in el or "copyright" in el or "all rights reserved" in el: 
            copyright_element.append(name_cleaner(el))
    if len(copyright_element)!=0:
        copyright_element = [name_cleaner(name) for name in copyright_element]
        return [name for name in copyright_element if name is not None]
    else:
        for element in tree.xpath('//text()'):
            el = element.lower()
            if el is None: continue
            el = el.replace("\n","").replace("\r","")
            el = re.sub(' +', ' ',el).lower()
            if el=="None" or el=="" or el==" ": continue
            if "©" in el or "copyright" in el: 
                for repl in copyright_replace: el = el.replace(repl,"")
                copyright_element.append(name_cleaner(el))
        copyright_element = list(dict.fromkeys(copyright_element))
        if len(copyright_element)==1: 
            cleaned_name = name_cleaner(copyright_element[0])
            if cleaned_name is not None: 
                return [cleaned_name]
            else: return []
        else: return []



def imprint_analyzer(domain, link, imprint_queue):
    """
    Loads the imprint URL and returns possible legal names that occur in it
    """
    # Maybe check for adresses (through city/country recognition and highlight the 
    # elements preceding the adress)
    # Expand list of legal entities in cleanco (through excel list downloaded)
    elements, imprint_names = [], []
    if urlparse(link)[1]=="": end_link = domain + "/" + link
    elif domain in link: end_link = link
    else:
        imprint_queue.put([])
        imprint_queue.task_done()
        return None
    if "http://" not in end_link and "https://" not in end_link: 
        end_link = "http://" + end_link
    try:
        response = urllib.request.urlopen(Request(end_link,headers={'User-Agent':User_ag}),
                                                       context = context, timeout=10)
        pageContent = response.read()
#        pageContent = clean_html(pageContent)
    except Exception as e: imprint_queue.put([]); return []
    tree = html.fromstring(pageContent)
    titles = [el.text for el in tree.xpath("//*") if el.tag=="title"]

    tree = tree.xpath("//text()")
    # get title element - if text is equal to title --> half the similarity ratio
    for el in tree:
        if el is None: continue
        el = el.replace("\n","").replace("\r","").replace("\t","").replace("\\t","")
        el = re.sub(' +', ' ',el).lower()
        if el=="None" or el=="" or el==" ": continue
        elements.append(el)
    for el in elements:
        name = cleanco(el)
        if name.type() is not None:
            if len(el)>50: continue
            imprint_names.append(el)
    try: imprint_names.extend(entity_search(pageContent, response))
    except: print(response.headers.get_content_charset())
    imprint_queue.put((imprint_names,titles))
    try:
        imprint_queue.task_done()
    except:
        imprint_queue.put([])
        imprint_queue.task_done()
    return []
        
def crawl_imprint(domain, soup_links):
    """ 
    Crawl for the imprint link (also: privacy statemant and disclaimer) and extract the 
    entity where we can find a legal name and which occurs most often --> test it against URL
    """
    imprint_links, imprint_threads, imprint_results = [], [], []
    for link in soup_links:
        try: aa = link.contents[0].lower(); aa = link["href"].lower(); 
        except: continue
        if domain in link["href"] or urlparse(link["href"])[1]=="": 
            if imprint_links in imprint_links: continue
            for word in imprint_trigger_words:
                if word in link["href"].lower() or word in link.contents[0].lower():
                    if link["href"].lower()=="#": pass
                    imprint_links.append(link["href"].lower())
    imprint_queue = queue.Queue()
    imprint_links = list(dict.fromkeys(imprint_links))
    for link in imprint_links:
        t = threading.Thread(target = imprint_analyzer, args=(domain, link, imprint_queue))
        t.daemon = True
        t.start()
        imprint_threads.append(t)
        time.sleep(0.1)
    counter, titles = 0, []
    while True:
        if imprint_queue.qsize()==len(imprint_threads): break
        if counter>20: break
        counter+=1
        time.sleep(0.2)
    for t in imprint_threads:
        try: res, title = imprint_queue.get(False)
        except: continue
        if res!=[]: 
            for imp in res: imprint_results.append(name_cleaner(imp))    
        if title!=[]:
            for imp in title: titles.append(name_cleaner(imp))
    return imprint_results, titles
    
def footer_crawler(tree):
    tree = tree.xpath('//footer//text()')
    elements, imprint_names = [], [] 
    for element in tree:
        el = element
        if el is None: continue
        el = el.replace("\n","").replace("\r","").replace("\t","")
        el = re.sub(' +', ' ',el).lower()
        if el=="None" or el=="" or el==" ": continue
        elements.append(el)
    for el in elements:
        name = cleanco(el)
        # somehow doesnt recognize "nielen schuman b.v."
        if name.type() is not None or name.country() is not None :
            imprint_names.append(name_cleaner(el))
    return imprint_names

def crawl_facebook(soup_links):
    """
    Crawl name from facebook page if link is given
    """
    fb_link, fb_name = None, []
    for link in soup_links:
        try:
            if ".facebook." in link["href"].lower():
                fb_link = link["href"].lower()
                break
        except: continue
    if fb_link is not None:
        if fb_link.startswith("//"): fb_link = fb_link[2:]
        if "www." in fb_link:
            fb_link = fb_link[fb_link.find(".")+1:]
            fb_link = "https://mobile." + fb_link
        else:
#            fb_link = fb_link.replace("https://www.","https://mobile.")
#            fb_link = fb_link.replace("http://www.","https://mobile.")
            fb_link = fb_link.replace("https://","https://mobile.")
            fb_link = fb_link.replace("http://","https://mobile.")
        try: driver.get(fb_link)
        except: return []
        if "Content not found" in driver.title or "Contenu introuvable" in driver.title or 'Pagina niet gevonden' in driver.title or 'Inhoud niet gevonden' in driver.title:
            return []
        tree = html.fromstring(driver.page_source)
        fb_name = tree.xpath("//*[@id='u_0_7']//text()")
    if fb_name == []: return []
    
    cleaned_name = name_cleaner(fb_name[0])
    if cleaned_name is None: return [fb_name[0]]
    else: return [cleaned_name]
       
def crawl_linkedin_link(soup_links):
    """
    Use name from linkedin link to cross-compare which results 
    are good (in case url is an abbreviation)
    """
    linkedin_name, linkedin_link = None, None
    for link in soup_links:
        try:
            if ".linkedin." in link["href"].lower() and "/company/" in link["href"].lower():
                linkedin_link = link["href"].lower()
                break
        except: continue
    if linkedin_link is not None:
        if linkedin_link[-1]=="/": linkedin_link=linkedin_link[:-1]
        linkedin_name_pre = linkedin_link[linkedin_link.find("/company/")+9:]
        splitted_num = linkedin_name_pre.find("/")
        if splitted_num<0: splitted_num = len(linkedin_name_pre)
        linkedin_name = linkedin_name_pre[:splitted_num]
        linkedin_name = linkedin_name.replace("-"," ")
        linkedin_name = re.sub("_\d+","",linkedin_name)
        if len(linkedin_name)<2: 
            return []
        if "trk=" in linkedin_name:
            trk = linkedin_name.find("trk=")
            trk_end = linkedin_name[trk:].find("?")
            if trk_end==-1: trk_end = len(linkedin_name)
            trk_arg = linkedin_name[trk:trk_end]
            linkedin_name = linkedin_name.replace(trk_arg,"")
    else: return []
    try: linkedin_link = int(linkedin_name); linkedin_name = None
    except: pass
    if linkedin_name is None: return []
    cleaned_name = name_cleaner(linkedin_name)
    if cleaned_name is None: return [linkedin_name]
    else: return [cleaned_name]

def split_url(domain, find_words=True):
    full_domain = domain[:domain.find(".")]
    splitted_url = " ".join(full_domain.split("-"))
    if find_words:
        for word in eng_words:
            if word in splitted_url:
                splitted_url = splitted_url.replace(word," "+word.upper()+" ")
    return [splitted_url.lower().strip()]
    
title_forbidded = ["\\n","\n","\\r","\r","\\t","\t","-","|"," home ","homepage"," site ","welcome"]

def get_site_title(used_selenium, tree, driver):
    """
    if none of the words in the imprint company name is found in URL:
        try page Title (remove "Home","Homepage","imprint" etc..)
    """
    site_title = None
    if used_selenium:
        site_title = driver.title
    else:
        site_title = tree.findtext('.//title')
    if site_title is None:
        pat = re.compile("<title>(.*?)<\/title>",re.DOTALL|re.M)
        pat1 = re.compile(".*(INTERESTING.*?)bodystuff",re.DOTALL|re.M)
        site_title = pat.findall(str(pageContent))[0]
    if site_title is not None:
        site_title = " "+site_title.lower()+" "
        for word in title_forbidded:
            site_title = site_title.replace(word," ")
        site_title = re.sub(' +', ' ',site_title).strip()
    else: site_title = ""
    return site_title

def entity_search(pageContent, response, title=""):
    """
    get entities in pagecontent that are recognized as legal entities
    """
    try:
        body = h.handle(pageContent)
    except:
        try: 
            charset = response.headers.get_content_charset()
            if charset is None: raise ValueError
        except: charset = 'utf-8'
        try: 
            content = pageContent.decode(charset)
            body = h.handle(content)
        except Exception as e: print(e,pageContent); return []
    body = body.replace("\n"," ").replace("\r"," ").replace("*"," ").replace("#"," ")
    entities = []
    doc = nlp(body)
    for ent in doc.ents:
        name = cleanco(str(ent))
        if name.type() is not None or name.country() is not None :
            name_to_add = name_cleaner(str(ent))
            if name_to_add is not None:
                entities.append(name_to_add)
        elif str(ent).lower() in title:
            name_to_add = name_cleaner(str(ent))
            if name_to_add is not None:
                entities.append(name_cleaner(str(name_to_add)))
    return list(dict.fromkeys(entities))

def bonus_calc(text, site_title):
    bonus = 0
    for word in text.lower().split():
        if word in site_title: bonus+=0.2
    return bonus

def check_warning(text):
    # check if something comes after legal suffix --> to avoid including addresses
    text = name_cleaner(text).lower()
    legal_pos = len(text.split())
    pos, warning = 0, False
    for word in text.split():
        if word in legal_suffixes and word is not "uk": 
            legal_pos = pos
        pos+=1
    if len(text.split()) - legal_pos > 2: warning = True
    return warning, legal_pos

import ssl
import socket
context = ssl._create_unverified_context()

def get_ssl_issuer(hostname):
    ctx = ssl.create_default_context()
    sock = socket.socket()
    sock.settimeout(1.5)
    s = ctx.wrap_socket(sock, server_hostname=hostname)
    try:
        try: s.connect((hostname, 443))
        except: 
            host="www."+hostname
            sock = socket.socket()
            sock.settimeout(1.5)
            s = ctx.wrap_socket(sock, server_hostname=host)
            s.connect((host, 443))
        cert = s.getpeercert()
        subject = dict(x[0] for x in cert['subject'])
        org = subject['organizationName']
        if "cloudflare" not in org.lower():
            return [org]
        else: 
            print("Cloudfare customer")
            return []
    except Exception as e:
        print(e)
        return []

from difflib import SequenceMatcher

def common_titles(titles):
    if len(titles)==0: return None
    str1 = titles[0]
    for title in titles:
        str2 = title
        seqMatch = SequenceMatcher(None,str1,str2)
        match = seqMatch.find_longest_match(0, len(str1), 0, len(str2))
        if (match.size!=0):
            str1 = str1[match.a: match.a + match.size]
        else: return None
    tt = []
    t = str1.split("-")
    for x in t: 
        if len(x)>2: tt.append(x)
    return " ".join(tt)

def calc_ratio(s,full_domain):
    try:
        abbr = ''.join(i[0] for i in s.lower().split())
        abbr_ratio = difflib.SequenceMatcher(None,abbr.lower(),full_domain.lower()).ratio()
        name_ratio = difflib.SequenceMatcher(None,s.lower(),full_domain.lower()).ratio()
        ratio = max(abbr_ratio,name_ratio)
        return ratio
    except:
        return 0.0001

def words_in_url(res, full_domain):
    activated = 0
    for words in res.lower().split():
        if words in full_domain.lower(): activated+=1
    return activated 

# Crawl Twitter
def clean_url(url):
    url = url.replace("http://","").replace("https://","")
    if "/" in url: url=url[:url.find("/")]
    url_tld = get_tld(url, fix_protocol=True).lower()
    url_main = url.replace("."+url_tld,"")
    if "." not in url_main: subdomain="www"
    else: subdomain = url_main[:url_main.find(".")]
    url_main = url_main[url_main.rfind(".")+1:]
    return "http://"+subdomain+"."+url_main+"."+url_tld

import io
try:
    from PIL import Image
except ImportError:
    import Image
import pytesseract

def logo_analyzer(domain):
    response = requests.get('https://logo.clearbit.com/'+domain)
    img = Image.open(io.BytesIO(response.content))
    return pytesseract.image_to_string(img)

def logo_crawl(tree):
    imgs = tree.xpath('//img')
    result_alts = []
    for img in imgs:
        if "logo" in img.get('src',"").lower() or "logo" in img.get('id',"") or "logo" in img.get("name",""):
            if img.attrib.get("alt")!="": 
                result_alts.append(img.attrib.get("alt"))
    return result_alts

def initial_loader(query_domain):
    response = ""
    domain = get_fld(query_domain, fix_protocol=True).lower()
    used_selenium = False
    try:
        x1 = time.monotonic()
        try: response = urllib.request.urlopen(Request("http://"+domain,headers={'User-Agent':User_ag}),
                                               context = context, timeout=5)
        except Exception as e:
            print(e)
            if "urlopen error timed out" in str(e): 
                print("we timed out after 10sec")
#                return name_cleaner(split_url(domain)[0], legal_clean=True)
            response = urllib.request.urlopen(Request("http://www"+domain,headers={'User-Agent': User_ag}),
                                               context = context, timeout=5)
            domain = "www."+domain
        pageContent = response.read()
        body = h.handle(str(pageContent))
        body = body.replace("\n","").replace("\\n","").replace("\\r","").replace("\\t","")
        if str(pageContent).count("\\")/len(pageContent)>0.4:
            raise ValueError('Probably ASCII Scripting')
        if len(str(body).split())<20: raise ValueError('Needs JS crawling')
        if len(body.split())<100 and "javascript" in body.lower() or "sorry" in body.lower(): 
            raise ValueError('Needs JS crawling')
        if len(body.split())<200 and "error" in body.lower() and "sorry" in body.lower(): 
            raise ValueError('Possible error')
        try: domain = get_fld(response.geturl(), fix_protocol=True).lower()
        except: pass
    except:
        used_selenium = True
        x1 = time.monotonic()
        try: 
            driver.get("http://"+domain)
            pageContent = driver.page_source
            site_title = driver.title
        except: 
            pageContent = ""
            site_title = "HTTP Error 404"
        body = h.handle(str(pageContent))
        body = body.replace("\n","").replace("\\n","").replace("\\r","").replace("\\t","")
        if site_title=="Not Found" or "HTTP Error 404" in str(pageContent) or len(body.split())<20:           
            try: 
                domain = "www."+domain
                driver.get("http://"+domain)
                pageContent = driver.page_source
                site_title = driver.title
            except: 
                print("website impossible to access")
                return name_cleaner(split_url(domain.replace("www.",""))[0], legal_clean=True)
        try: 
            domain = get_fld(driver.current_url, fix_protocol=True).lower()
        except: pass
        x2 = time.monotonic()
        print(x2-x1)
#    pageContent = clean_html(pageContent)
#    print("domain")
    ssl_issuer = get_ssl_issuer(domain)
    tree = html.fromstring(pageContent)
    
    site_title = get_site_title(used_selenium, tree, driver)
    soup = BeautifulSoup(pageContent, 'html.parser')
    soup_links = soup.find_all('a')
    
    full_domain = domain[:domain.find(".")]
    # Class 1
    copyright_content = copyright_crawler(tree)
    footer_content = footer_crawler(tree)
    linkedin_name = crawl_linkedin_link(soup_links)
    print("copyright:",copyright_content)
    print("footer legal:",footer_content)
    print("linkedin_name:",linkedin_name)
    print("SSL issuer:",ssl_issuer)
   
    for s in ssl_issuer:
        s = name_cleaner(s,legal_clean=True)
        ratio = calc_ratio(s,full_domain)
        if ratio>0.6: return name_cleaner(s, legal_clean=True)
        
    # if other results equals to linkedin results when concatenated, delete linkedin results
    facebook_name = crawl_facebook(soup_links)
    print("facebook_name:",facebook_name)
    if facebook_name!=[]: 
        return name_cleaner(facebook_name[0], legal_clean=True)

    step1_results = copyright_content + footer_content + ssl_issuer + facebook_name
    use_linkedin = True
    if len(linkedin_name)>0:
        if calc_ratio(linkedin_name[0],full_domain)<0.25: use_linkedin = False
        for res in step1_results:
            if name_cleaner(res,legal_clean=True).replace(" ","").lower() in linkedin_name[0].lower():
                use_linkedin = False
    if use_linkedin: step1_results+=linkedin_name

    if len(copyright_content+footer_content)==0:
        ent_search = entity_search(pageContent, response, site_title)
        step1_results+=ent_search
        print("Entity search",ent_search)
        
    res_list = []
    warning_count = 0 
    for res in step1_results:
        if res is None: continue
        ratio = difflib.SequenceMatcher(None,res.lower(),full_domain.lower()).ratio()
        ratio += bonus_calc(res, site_title.split())
        if ratio>0.25:
            adress_warning, legal_pos = check_warning(res)
            if "website" in res.lower() or "homepage" in res.lower():
                ratio = ratio /2
                warning_count += 1
            elif words_in_url(res, full_domain)==0 or len(res)>40 or adress_warning:
                ratio = ratio /2
                warning_count += 1                
            res_list.append((ratio,res))
    res_list.sort(reverse = True)
    if len(res_list)>0 and warning_count<len(res_list) and len(res_list)>1:
        return name_cleaner(res_list[0][1], legal_clean=True)
    
    # Class 3
    imprint_content, titles = crawl_imprint(domain, soup_links)
    print("imprint_name:",imprint_content)
    res_list, warning_count  = [], 0
    for res in imprint_content + step1_results:
        if res is None: continue
        ratio = difflib.SequenceMatcher(None,res.lower(),full_domain.lower()).ratio()
        if ratio>0.25:
            adress_warning, legal_pos = check_warning(res)
            if res in titles or "website" in res.lower() or "homepage" in res.lower() or adress_warning or len(res)>40:
                res = res.lower().replace("home - "," ").replace(" home "," ")
                ratio = ratio/2
                warning_count +=1
            res_list.append((ratio,res))
    titles = list(dict.fromkeys(titles))
    if len(titles)>2:
        common_title = name_cleaner(common_titles(titles))
        res_list.append((calc_ratio(common_title,full_domain),common_title))
        print("Common title:", common_title)
    print(res_list)
    res_list.sort(reverse = True)
    if len(res_list)>0 and warning_count<len(res_list): 
        return name_cleaner(res_list[0][1], legal_clean=True)
    
    # check overlapp with abbreviations tha we found before using split_url e.g. abf--> Associated British Foods Plc
    results = copyright_content+footer_content+linkedin_name+imprint_content
    res_list = []
    for res in results:
        if res is None: continue
        abbr = cleanco(res.lower()).clean_name()
        abbr = ''.join(i[0] for i in abbr.split())
        ratio = difflib.SequenceMatcher(None,abbr.lower(),full_domain.lower()).ratio()
        if ratio>0.25:
            adress_warning, legal_pos = check_warning(res)
            if adress_warning and legal_pos: res = " ".join(res.split()[:legal_pos+1])
            if res in titles:
                res = res.lower().replace("home - "," ").replace("home"," ")
                ratio = ratio/2
            res_list.append((ratio,res))
    res_list.sort(reverse = True)
    if len(res_list)>0: 
        return name_cleaner(res_list[0][1], legal_clean=True)

    # Class 4
    splited_url = split_url(domain)
    print("we used splitted url:",splited_url)
    return name_cleaner(splited_url[0], legal_clean=True)

    #difflib.SequenceMatcher(None,domain_w.lower(),res["title"].lower()).ratio()
    return  footer_content + copyright_content + imprint_content + facebook_name + linkedin_name + splited_url

from openpyxl import Workbook, load_workbook

def excel_test():
    wb = load_workbook("dtn_perf.xlsx")
    ws = wb.active
    x = 1
    x1 = time.monotonic()
    while True:
        x += 1
        print(x)
        query_domain = ws.cell(row = x, column = 1).value
        if query_domain is None or query_domain=="None" or query_domain=="": break
        print(query_domain)
        try: 
            found_name = initial_loader(query_domain)
            ws.cell(row = x, column = 2).value = found_name
        except Exception as e:
            print(e)
        print(found_name)
    x2 = time.monotonic()
    print("time taken:",x2-x1)
    wb.save("dtn_perf.xlsx")
    
        
        